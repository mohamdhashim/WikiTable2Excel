import requests  # to get HTML through Request from the WikiPage
from bs4 import BeautifulSoup  # used to extract data from HTML
from openpyxl import Workbook  # to manipulating our tables
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import re


links = {}  # global Var that Store all linkes related to table


def clean_row(full_row, lan):
    '''
        function takes a table row ('tr') and convert it to a cleaned List use it in the Sheet
        Output: list of cleaned_data(ready to use)
    '''
    row = full_row.findChildren('th') + full_row.findChildren('td')
    clean_row = []

    for column in row:
        try:
            # convert numbers to int (to avoid bugs of int operations on strings in excel ex: spaces, lines after nums)
            text = int(column.get_text())
        except:
            # Regex to remove special chars from text
            text = re.sub(r'[\W_]+', ' ', column.get_text())

        clean_row.append(text)

        link = column.find('a')
        if link:
            links[text] = f'https://{lan}.wikipedia.org' + link['href']
        else:
            links[text] = None

    return clean_row


def get_language_of_wiki_table(URL):
    '''
        Return The language of Wikipedia based on wiki URL standards ==> https://en.wikipedia.org/wiki/URL (more info)
        ex : 'https://ar.wikipedia.org/wiki/الحرب_الفرنسية_البروسية' ==> 'ar'
        ex : 'https://de.wikipedia.org/wiki/Deutsch-Französischer_Krieg' ==> 'de'
        ex : 'https:arz.wikipedia.org/wiki/ويكيبيديا_مصرى
    '''

    if(URL[10] == '.'):  # if language is 2 chars ['ar','fr','du']
        return URL[8:10]
    elif(URL[11] == '.'):  # if language is 3 chars ['arz']
        return URL[8:11]
    else:
        return 'en'


def scrape_wiki_table(url, save_to_dir = 'output.xlsx', attribute='class', attr_value='wikitable', order=1):
    # attribute: it's used to get a specific table in the page by it's html attribute ex: 'class', 'id'
    # arr_value: it's the value of the Html arrtibute ex: <classname> , <id value>
    # order: to specifiy the table in the page by it's order in page tables count from 1

    global links
    links = {}
    lan = get_language_of_wiki_table(url)  # lan == language use in the link

    page_request = requests.get(url).text  # 1.Get Html
    
    # 2.Parse it using BeautifulSoup
    page_html = BeautifulSoup(page_request, 'html.parser')

    # 3.find a table by {'class':'wikitable'}
    table = page_html.find_all('table', {attribute: attr_value})
    table = table[order-1]

    table_body = table.find('tbody', recursive=False)  # 4.Get Table Body
    rows = table_body.findChildren('tr')  # 5. Get each row in the table

    wb = Workbook()  # Create New Excel WorkBook
    ws = wb.active  # Create New Sheet in Current WorkBook

    for row in rows[0:]:
        ws.append(clean_row(row, lan))

    table = Table(displayName="Table1", ref="A1:" +
                  get_column_letter(ws.max_column) + str(ws.max_row))
    ws.add_table(table)  # Create Table for our Current range

    # adding Hyperlinks to each table Cell if Exists using links{}
    for i in range(1, ws.max_row+1):
        for j in range(1, ws.max_column+1):
            link = links.get(ws.cell(row=i, column=j).value, None)
            ws.cell(row=i, column=j).hyperlink = link

    wb.save(save_to_dir)  # dir must end with .xlsx



if __name__ == '__main__':
    
    ##Exmple UseCase to get 2nd table in URL
    
    url = 'https://en.wikipedia.org/wiki/List_of_European_Cup_and_UEFA_Champions_League_finals'
    output = 'UEFA_Champions.xlsx'
    scrape_wiki_table(url = url,save_to_dir=output, order=2)
    
    
    url = 'https://ar.wikipedia.org/wiki/قائمة_أفضل_مئة_رواية_عربية'
    output = 'book_list.xlsx'
    scrape_wiki_table(url = url,save_to_dir=output)
